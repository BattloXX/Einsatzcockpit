from datetime import datetime, timedelta
from io import BytesIO

from sqlalchemy import BigInteger, create_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker

from app.core.tenant import set_tenant_context
from app.db import Base
from app.models.incident import Incident, IncidentColumn, IncidentVehicle
from app.models.master import AlarmType, FireDept, VehicleMaster
from app.services.stats_service import get_stats


@compiles(BigInteger, "sqlite")
def _bigint_sqlite(element, compiler, **kw):
    return "INTEGER"


def test_stats_aggregates_and_reaction_time():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    set_tenant_context(db, None)
    org_a = FireDept(slug="stats-service-a", name="Stats A", timezone="Europe/Vienna")
    org_b = FireDept(slug="stats-service-b", name="Stats B", timezone="Europe/Vienna")
    db.add_all([org_a, org_b])
    db.flush()
    db.add_all([
        AlarmType(org_id=org_a.id, code="F1", category="F", label="Brand"),
        AlarmType(org_id=org_a.id, code="T1", category="T", label="Technisch"),
    ])
    started = datetime(2026, 3, 10, 8, 0)
    fire = Incident(primary_org_id=org_a.id, alarm_type_code="F1", started_at=started,
                    closed_at=started + timedelta(minutes=60), is_exercise=False)
    technical = Incident(primary_org_id=org_a.id, alarm_type_code="T1", started_at=started,
                          is_exercise=False)
    exercise = Incident(primary_org_id=org_a.id, alarm_type_code="B1", started_at=started,
                        is_exercise=True)
    outside = Incident(primary_org_id=org_a.id, alarm_type_code="B1",
                       started_at=datetime(2025, 3, 10), is_exercise=False)
    foreign = Incident(primary_org_id=org_b.id, alarm_type_code="B1", started_at=started,
                       is_exercise=False)
    db.add_all([fire, technical, exercise, outside, foreign])
    db.flush()
    column = IncidentColumn(incident_id=fire.id, code="active", title="Aktiv", column_kind="vehicles")
    vehicle = VehicleMaster(dept_id=org_a.id, code="TLF", name="TLF")
    db.add_all([column, vehicle])
    db.flush()
    db.add(IncidentVehicle(incident_id=fire.id, column_id=column.id, vehicle_master_id=vehicle.id,
                           created_at=started + timedelta(minutes=7)))
    db.commit()

    stats = get_stats(db, org_a.id, started.date(), started.date(), user=None)
    assert stats.total == 2
    assert stats.total_exercises == 1
    assert stats.fire_count == 1
    assert stats.technical_count == 1
    assert stats.other_count == 0
    assert stats.avg_duration_min == 60
    assert stats.avg_time_to_first_vehicle_min == 7
    assert sum(row["count"] for row in stats.by_alarm_type) == stats.total
    assert foreign.id not in [row.id for row in stats.incidents]
    assert outside.id not in [row.id for row in stats.incidents]

    from openpyxl import load_workbook

    from app.services.excel_export_service import exportiere_einsatzstatistik
    from app.services.pdf_service import render_statistik_bericht_pdf

    xlsx = exportiere_einsatzstatistik(stats, started.date(), started.date(), org_a)
    workbook = load_workbook(BytesIO(xlsx))
    assert workbook["Einsaetze"].max_row == stats.total + 1
    pdf = render_statistik_bericht_pdf(stats, org_a, started.date(), started.date())
    assert pdf.startswith(b"%PDF")
    db.close()


def test_stats_timestamp_kpis_use_median_for_delayed_records():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    set_tenant_context(db, None)
    org = FireDept(slug="stats-median", name="Stats Median", timezone="Europe/Vienna")
    db.add(org)
    db.flush()
    db.add(AlarmType(org_id=org.id, code="T1", category="T", label="Technisch"))

    started = datetime(2026, 3, 10, 8, 0)
    for index, (duration, dispatch) in enumerate([(30, 5), (60, 10), (6000, 5000)]):
        incident = Incident(
            primary_org_id=org.id, alarm_type_code="T1",
            started_at=started + timedelta(minutes=index),
            closed_at=started + timedelta(minutes=index + duration), is_exercise=False,
        )
        db.add(incident)
        db.flush()
        column = IncidentColumn(
            incident_id=incident.id, code="active", title="Aktiv", column_kind="vehicles",
        )
        vehicle = VehicleMaster(dept_id=org.id, code=f"FZG{index}", name=f"Fahrzeug {index}")
        db.add_all([column, vehicle])
        db.flush()
        db.add(IncidentVehicle(
            incident_id=incident.id, column_id=column.id, vehicle_master_id=vehicle.id,
            created_at=incident.started_at + timedelta(minutes=dispatch),
        ))
    db.commit()

    stats = get_stats(db, org.id, started.date(), started.date(), user=None)
    assert stats.avg_duration_min == 60
    assert stats.avg_time_to_first_vehicle_min == 10
    db.close()


def test_stats_alarm_to_scene_uses_median_and_omits_missing_values():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    set_tenant_context(db, None)
    org = FireDept(
        slug="stats-on-scene", name="Stats Einsatzort", timezone="Europe/Vienna"
    )
    db.add(org)
    db.flush()
    db.add(AlarmType(org_id=org.id, code="T1", category="T", label="Technisch"))

    started = datetime(2026, 3, 10, 8, 0)
    db.add_all([
        Incident(
            primary_org_id=org.id,
            alarm_type_code="T1",
            started_at=started,
            on_scene_at=started + timedelta(minutes=5),
            is_exercise=False,
        ),
        Incident(
            primary_org_id=org.id,
            alarm_type_code="T1",
            started_at=started + timedelta(minutes=1),
            on_scene_at=started + timedelta(minutes=16),
            is_exercise=False,
        ),
        Incident(
            primary_org_id=org.id,
            alarm_type_code="T1",
            started_at=started + timedelta(minutes=2),
            on_scene_at=None,
            is_exercise=False,
        ),
        Incident(
            primary_org_id=org.id,
            alarm_type_code="T1",
            started_at=started + timedelta(days=1),
            on_scene_at=None,
            is_exercise=False,
        ),
    ])
    db.commit()

    stats = get_stats(db, org.id, started.date(), started.date(), user=None)
    assert stats.median_alarm_to_scene_min == 10

    day_without_on_scene = started.date() + timedelta(days=1)
    assert (
        get_stats(db, org.id, day_without_on_scene, day_without_on_scene, user=None)
        .median_alarm_to_scene_min
        is None
    )
    db.close()


def test_stats_map_markers_omit_invalid_coordinates_for_twelve_month_range():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    set_tenant_context(db, None)
    org = FireDept(slug="stats-map-coordinates", name="Stats Map", timezone="Europe/Vienna")
    db.add(org)
    db.flush()

    started = datetime(2025, 9, 10, 8, 0)
    coordinates = [
        (47.5, 14.5),
        (47.5, 14.5),  # Duplikate sind gueltige, getrennte Einsatzorte.
        (-33.9, 151.2),  # Weit entfernte, aber gueltige WGS84-Koordinate.
        (None, 14.5),
        (47.5, None),
        (0.0, 0.0),
        (float("inf"), 14.5),
        (47.5, float("-inf")),
        (90.1, 14.5),
        (47.5, 180.1),
    ]
    for index, (lat, lng) in enumerate(coordinates):
        db.add(Incident(
            primary_org_id=org.id,
            alarm_type_code="T1",
            started_at=started + timedelta(days=index * 30),
            is_exercise=False,
            lat=lat,
            lng=lng,
        ))
    db.commit()

    stats = get_stats(
        db,
        org.id,
        started.date(),
        (started + timedelta(days=365)).date(),
        user=None,
    )

    assert [(marker["lat"], marker["lng"]) for marker in stats.map_markers] == [
        (-33.9, 151.2),
        (47.5, 14.5),
        (47.5, 14.5),
    ]
    db.close()
