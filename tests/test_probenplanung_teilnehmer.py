"""Routentests fuer Phase 8: schnelle, regressionsfreie Teilnehmererfassung."""

from io import BytesIO
from unittest.mock import patch

import openpyxl

from app.core.tenant import set_tenant_context
from app.db import SessionLocal
from app.models.incident import Incident
from app.models.master import Member, OrgSettings
from app.models.teilnahme import Teilnahme, TeilnahmeStatus
from tests.test_probenplanung_checkliste import ORG_ID, _flags, _login, _probe_anlegen, _probeart, _user


def _member(suffix: str, *, active: bool = True) -> int:
    db = SessionLocal()
    set_tenant_context(db, None)
    try:
        row = Member(org_id=ORG_ID, firstname=f"Phase8{suffix}", lastname="Teilnehmer", active=active)
        db.add(row)
        db.commit()
        return row.id
    finally:
        db.close()


def _probe(client, username: str) -> tuple[str, int]:
    _user(username, "probenverwalter")
    _flags()
    csrf = _login(client, username)
    probeart_id = _probeart(0, f"Phase8 Art {username}")
    return csrf, _probe_anlegen(client, csrf, probeart_id, f"Phase8 Probe {username}")


def _row(termin_id: int, member_id: int) -> Teilnahme | None:
    db = SessionLocal()
    set_tenant_context(db, None)
    try:
        return db.query(Teilnahme).filter_by(bezug_typ="uebung", bezug_id=termin_id, mitglied_id=member_id).first()
    finally:
        db.close()


def test_default_bleibt_nicht_erfasst_und_alle_aktiven_werden_gelistet(client):
    csrf, termin_id = _probe(client, "phase8_default")
    del csrf
    active_id = _member("DefaultAktiv")
    inactive_id = _member("DefaultInaktiv", active=False)

    response = client.get(f"/probenplanung/{termin_id}/teilnehmer")
    assert response.status_code == 200
    assert "Phase8DefaultAktiv Teilnehmer" in response.text
    assert "Phase8DefaultInaktiv Teilnehmer" not in response.text
    assert "noch nicht erfasst" in response.text
    assert "unentschuldigt" in response.text
    assert _row(termin_id, active_id) is None
    assert _row(termin_id, inactive_id) is None


def test_status_patch_summary_und_altfelder_sind_konsistent(client):
    csrf, termin_id = _probe(client, "phase8_status")
    member_ids = [_member(f"Status{i}") for i in range(3)]
    expected = [
        ("anwesend", True, False),
        ("entschuldigt", False, True),
        ("unentschuldigt", False, False),
    ]
    for member_id, (status, _, _) in zip(member_ids, expected, strict=True):
        response = client.patch(
            f"/probenplanung/{termin_id}/teilnehmer/{member_id}",
            data={"_csrf": csrf, "status": status, "notiz": "Erfasst"},
        )
        assert response.status_code == 200

    db = SessionLocal()
    set_tenant_context(db, None)
    try:
        active_count = db.query(Member).filter(Member.org_id == ORG_ID, Member.active.is_(True)).count()
        rows = {
            row.mitglied_id: row
            for row in db.query(Teilnahme).filter(Teilnahme.bezug_id == termin_id).all()
        }
        for member_id, (status, ausgerueckt, entschuldigt) in zip(member_ids, expected, strict=True):
            assert (rows[member_id].status, rows[member_id].ausgerueckt, rows[member_id].entschuldigt) == (
                status,
                ausgerueckt,
                entschuldigt,
            )
    finally:
        db.close()
    assert "1 anwesend" in response.text
    assert "1 entschuldigt" in response.text
    assert "1 unentschuldigt" in response.text
    assert f"{active_count - 3} noch nicht erfasst" in response.text


def test_alle_anwesend_aendert_nur_nicht_erfasste_und_zuruecksetzen(client):
    csrf, termin_id = _probe(client, "phase8_bulk")
    present_id, excused_id, absent_id, open_id = [_member(f"Bulk{i}") for i in range(4)]
    for member_id, status in (
        (present_id, "anwesend"),
        (excused_id, "entschuldigt"),
        (absent_id, "unentschuldigt"),
    ):
        client.patch(
            f"/probenplanung/{termin_id}/teilnehmer/{member_id}",
            data={"_csrf": csrf, "status": status},
        )

    response = client.post(
        f"/probenplanung/{termin_id}/teilnehmer/alle-anwesend", data={"_csrf": csrf}
    )
    assert response.status_code == 200
    assert _row(termin_id, present_id).status == "anwesend"  # type: ignore[union-attr]
    assert _row(termin_id, excused_id).status == "entschuldigt"  # type: ignore[union-attr]
    assert _row(termin_id, absent_id).status == "unentschuldigt"  # type: ignore[union-attr]
    assert _row(termin_id, open_id).status == "anwesend"  # type: ignore[union-attr]
    assert "0 noch nicht erfasst" in response.text

    reset = client.post(f"/probenplanung/{termin_id}/teilnehmer/zuruecksetzen", data={"_csrf": csrf})
    assert reset.status_code == 200
    assert "0 unentschuldigt" in reset.text
    for member_id in (present_id, excused_id, absent_id, open_id):
        row = _row(termin_id, member_id)
        assert row is not None
        assert (row.status, row.ausgerueckt, row.entschuldigt) == ("nicht_erfasst", False, False)


def test_einsatz_mannschaft_pdf_und_xlsx_lesen_weiterhin_altfelder(client):
    _user("phase8_regression", "incident_leader")
    _flags()
    _login(client, "phase8_regression")
    member_id = _member("Regression")
    db = SessionLocal()
    set_tenant_context(db, None)
    try:
        incident = Incident(primary_org_id=ORG_ID, alarm_type_code="T1", status="active")
        db.add(incident)
        db.flush()
        row = Teilnahme(org_id=ORG_ID, bezug_typ="einsatz", bezug_id=incident.id, mitglied_id=member_id)
        row.set_status(TeilnahmeStatus.ANWESEND)
        db.add(row)
        db.commit()
        incident_id = incident.id
    finally:
        db.close()

    page = client.get(f"/einsatz/{incident_id}/mannschaft")
    assert page.status_code == 200
    assert "Phase8Regression Teilnehmer" in page.text
    assert 'title="Ausgerueckt"' in page.text
    assert "checked" in page.text

    captured: dict = {}

    def fake_pdf(**kwargs):
        captured.update(kwargs)
        return b"%PDF-regression"

    with patch("app.services.pdf_service.render_teilnahme_pdf", side_effect=fake_pdf):
        pdf = client.get(f"/teilnahme/einsatz/{incident_id}/export.pdf")
    assert pdf.status_code == 200 and pdf.content == b"%PDF-regression"
    exported = captured["teilnahmen"]
    assert [(item.anzeige_name, item.ausgerueckt, item.entschuldigt) for item in exported] == [
        ("Phase8Regression Teilnehmer", True, False)
    ]

    xlsx = client.get(f"/teilnahme/einsatz/{incident_id}/export.xlsx")
    assert xlsx.status_code == 200
    sheet = openpyxl.load_workbook(BytesIO(xlsx.content)).active
    assert [sheet.cell(1, column).value for column in range(1, 6)] == [
        "Nr.", "Name", "Aktiv", "Funktion", "Fahrzeug"
    ]
    assert [sheet.cell(2, column).value for column in range(1, 4)] == [
        1, "Phase8Regression Teilnehmer", "✓"
    ]


def test_rollen_und_modulguards_auf_teilnehmerrouten(client):
    _, termin_id = _probe(client, "phase8_guards_editor")
    member_id = _member("Guard")
    _user("phase8_guards_reader", "readonly")
    reader_csrf = _login(client, "phase8_guards_reader")
    assert client.get(f"/probenplanung/{termin_id}/teilnehmer").status_code == 200
    assert client.patch(
        f"/probenplanung/{termin_id}/teilnehmer/{member_id}",
        data={"_csrf": reader_csrf, "status": "anwesend"},
    ).status_code == 403

    editor_csrf = _login(client, "phase8_guards_editor")
    db = SessionLocal()
    set_tenant_context(db, None)
    try:
        db.query(OrgSettings).filter(OrgSettings.org_id == ORG_ID).one().probenplanung_modul_aktiv = False
        db.commit()
    finally:
        db.close()
    assert client.get(f"/probenplanung/{termin_id}/teilnehmer").status_code == 404
    assert client.post(
        f"/probenplanung/{termin_id}/teilnehmer/alle-anwesend", data={"_csrf": editor_csrf}
    ).status_code == 404
