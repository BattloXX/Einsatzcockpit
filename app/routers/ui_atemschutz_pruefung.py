"""Atemschutzgeräteprüfung – öffentliches Erfassungsformular (Token/QR) + Liste/Export.

Struktur analog app/routers/ui_fahrtenbuch.py (öffentlicher Token-Zugriff ohne Login)
kombiniert mit app/routers/ui_termin.py (angemeldete Liste + Exporte).
"""
from __future__ import annotations

import io
import logging
from datetime import UTC, date, datetime, timedelta

from fastapi import APIRouter, BackgroundTasks, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.core.audit import write_incident_change
from app.core.templating import templates
from app.core.timezones import format_local_datetime
from app.db import get_db
from app.models.atemschutz_pruefung import AtemschutzGeraet, AtemschutzPruefung
from app.models.incident import Incident
from app.models.master import Member, MemberQualification, OrgSettings, Qualification
from app.services.atemschutz_pruefung_service import melde_defekt_background

router = APIRouter()
logger = logging.getLogger("einsatzleiter.atemschutz_pruefung")


def _require_login(request: Request):
    user = getattr(request.state, "user", None)
    if user is None:
        raise HTTPException(status_code=401, detail="Nicht angemeldet")
    return user


def _resolve_org_by_token(token: str, db: Session) -> OrgSettings | None:
    """Liefert OrgSettings zur Token-Org oder None (analog Fahrtenbuch-Token)."""
    return (
        db.query(OrgSettings)
        .filter(OrgSettings.atemschutz_pruef_token == token)
        .execution_options(include_all_tenants=True)
        .first()
    )


def _agt_members(org_id: int, db: Session) -> list[Member]:
    return (
        db.query(Member)
        .join(MemberQualification, MemberQualification.member_id == Member.id)
        .join(Qualification, Qualification.id == MemberQualification.qualification_id)
        .filter(Member.active.is_(True), Member.org_id == org_id, Qualification.code == "AGT")
        .execution_options(include_all_tenants=True)
        .order_by(Member.lastname, Member.firstname)
        .distinct()
        .all()
    )


def _aktive_geraete(org_id: int, db: Session) -> list[AtemschutzGeraet]:
    return (
        db.query(AtemschutzGeraet)
        .filter(AtemschutzGeraet.org_id == org_id, AtemschutzGeraet.aktiv.is_(True))
        .execution_options(include_all_tenants=True)
        .order_by(AtemschutzGeraet.nummer)
        .all()
    )


def _recent_incidents(org_id: int, db: Session) -> list[Incident]:
    grenze = datetime.now(UTC) - timedelta(days=3)
    return (
        db.query(Incident)
        .filter(Incident.primary_org_id == org_id, Incident.started_at >= grenze)
        .execution_options(include_all_tenants=True)
        .order_by(Incident.started_at.desc())
        .limit(20)
        .all()
    )


# ── Öffentlich (kein Login, token-basiert) ──────────────────────────────────

@router.get("/ap/{token}", response_class=HTMLResponse)
async def pruefung_formular(request: Request, token: str, db: Session = Depends(get_db)):
    org_settings = _resolve_org_by_token(token, db)
    if not org_settings or not org_settings.atemschutz_pruefung_modul_aktiv:
        raise HTTPException(status_code=404, detail="Token ungültig")
    return templates.TemplateResponse(request, "atemschutz_pruefung/formular.html", {
        "user": None,
        "org": org_settings,
        "token": token,
        "geraete": _aktive_geraete(org_settings.org_id, db),
        "agt_members": _agt_members(org_settings.org_id, db),
        "now": datetime.now(UTC),
        "form_daten": {},
        "fehler": None,
    })


@router.get("/ap/{token}/einsaetze", response_class=HTMLResponse)
async def pruefung_einsaetze_dropdown(request: Request, token: str, db: Session = Depends(get_db)):
    org_settings = _resolve_org_by_token(token, db)
    if not org_settings:
        return HTMLResponse("")
    incidents = _recent_incidents(org_settings.org_id, db)
    return templates.TemplateResponse(request, "atemschutz_pruefung/_einsatz_dropdown.html", {
        "user": None,
        "incidents": incidents,
    })


@router.post("/atemschutz-pruefung", response_class=HTMLResponse)
async def pruefung_speichern(
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    user = getattr(request.state, "user", None)
    form = await request.form()
    token = str(form.get("t", "") or "")
    org_settings = _resolve_org_by_token(token, db) if token else None
    if not user and not org_settings:
        raise HTTPException(status_code=401, detail="Nicht authentifiziert")

    org_id = user.org_id if user else org_settings.org_id  # type: ignore[union-attr]

    def _fehler_zurueck(text: str):
        return templates.TemplateResponse(request, "atemschutz_pruefung/formular.html", {
            "user": user,
            "org": org_settings,
            "token": token,
            "geraete": _aktive_geraete(org_id, db),
            "agt_members": _agt_members(org_id, db),
            "now": datetime.now(UTC),
            "form_daten": dict(form),
            "fehler": text,
        }, status_code=422)

    geraet_id = int(form.get("geraet_id") or 0)
    geraet = (
        db.query(AtemschutzGeraet)
        .filter(AtemschutzGeraet.id == geraet_id, AtemschutzGeraet.org_id == org_id)
        .execution_options(include_all_tenants=True)
        .first()
    )
    if not geraet:
        return _fehler_zurueck("Bitte ein gültiges Atemschutzgerät wählen.")

    traeger_member_id = form.get("traeger_member_id") or ""
    traeger_member = None
    if traeger_member_id:
        traeger_member = (
            db.query(Member)
            .filter(Member.id == int(traeger_member_id), Member.org_id == org_id)
            .execution_options(include_all_tenants=True)
            .first()
        )
    traeger_free_text = (form.get("traeger_free_text") or "").strip() or None
    if not traeger_member and not traeger_free_text:
        return _fehler_zurueck("Bitte einen Atemschutzträger wählen oder eintragen.")

    eingesetzt_am_raw = form.get("eingesetzt_am") or ""
    try:
        eingesetzt_am = date.fromisoformat(str(eingesetzt_am_raw))
    except ValueError:
        return _fehler_zurueck("Bitte ein gültiges Datum eingeben.")

    einsatz_art = str(form.get("einsatz_art") or "uebung")
    if einsatz_art not in ("uebung", "einsatz"):
        einsatz_art = "uebung"

    incident_id: int | None = None
    incident_id_raw = form.get("incident_id") or ""
    if einsatz_art == "einsatz" and incident_id_raw:
        inc = (
            db.query(Incident)
            .filter(Incident.id == int(incident_id_raw), Incident.primary_org_id == org_id)
            .execution_options(include_all_tenants=True)
            .first()
        )
        if inc:
            incident_id = inc.id

    def _pflicht_int(key: str) -> int | None:
        raw = form.get(key)
        if raw in (None, ""):
            return None
        try:
            return int(float(raw))  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None

    # Drucküberprüfung nach dem Gebrauch: 3 Pflichtwerte, i.O./NICHT i.O. wird
    # automatisch aus den Grenzwerten in AtemschutzPruefung berechnet (kein
    # manuelles Umschalten mehr fuer die Hochdruckpruefung noetig).
    flaschendruck_bar = _pflicht_int("flaschendruck_bar")
    druckabfall_bar = _pflicht_int("druckabfall_bar")
    rueckzugssignal_bar = _pflicht_int("rueckzugssignal_bar")
    if flaschendruck_bar is None or druckabfall_bar is None or rueckzugssignal_bar is None:
        return _fehler_zurueck("Bitte Flaschendruck, Hochdruck-Dichtprüfung und Warnsignal-Prüfung eintragen.")

    sichtpruefung_ok = form.get("sichtpruefung_ok") == "ok"
    geraet_einsatzbereit_ok = form.get("geraet_einsatzbereit_ok") == "ok"
    hochdruckpruefung_ok = druckabfall_bar <= AtemschutzPruefung.HOCHDRUCK_DICHTPRUEFUNG_MAX_BAR_MIN
    defekt_info = (form.get("defekt_info") or "").strip() or None

    pruefung = AtemschutzPruefung(
        org_id=org_id,
        # Bereits geladene Objekte direkt zuweisen (nicht nur *_id) — verhindert,
        # dass Jinja beim Rendern der Erfolgsseite die lazy="joined"-Relationship
        # nachlädt: für anonyme Token-Requests wurde nie ein Tenant-Kontext
        # gesetzt, ein Nachladen von atemschutz_geraet/member würde dort mit
        # TenantContextMissing abbrechen (siehe app/core/tenant.py).
        geraet=geraet,
        traeger_member=traeger_member,
        traeger_free_text=traeger_free_text,
        eingesetzt_am=eingesetzt_am,
        ort_text=(form.get("ort_text") or "").strip() or None,
        einsatz_art=einsatz_art,
        incident_id=incident_id,
        flasche_gewechselt=form.get("flasche_gewechselt") == "on",
        flaschendruck_bar=flaschendruck_bar,
        sichtpruefung_ok=sichtpruefung_ok,
        druckabfall_bar=druckabfall_bar,
        hochdruckpruefung_ok=hochdruckpruefung_ok,
        rueckzugssignal_bar=rueckzugssignal_bar,
        geraet_einsatzbereit_ok=geraet_einsatzbereit_ok,
        defekt_info=defekt_info,
        created_via="intern" if user else "public",
        created_by_user_id=user.id if user else None,
    )
    alles_ok = pruefung.alles_ok
    if not alles_ok and not defekt_info:
        return _fehler_zurueck("Bitte eine Zusatzinfo zum Defekt eintragen.")

    db.add(pruefung)
    db.flush()

    if incident_id:
        write_incident_change(
            db, incident_id, "as_pruefung.created", "atemschutz_pruefung", pruefung.id,
            before=None,
            after={
                "geraet_label": geraet.anzeige_label,
                "alles_ok": alles_ok,
                "defekte_punkte": ", ".join(pruefung.defekte_punkte),
            },
            user_id=user.id if user else None,
        )

    db.commit()

    if not alles_ok:
        base_url = str(request.base_url).rstrip("/")
        background_tasks.add_task(melde_defekt_background, pruefung.id, base_url)

    return templates.TemplateResponse(request, "atemschutz_pruefung/erfolg.html", {
        "user": user,
        "pruefung": pruefung,
        "token": token,
    })


# ── Angemeldet: Liste, Detail, Exporte (für jeden Nutzer lesbar) ────────────

def _pruefungen_query(org_id: int, db: Session):
    return (
        db.query(AtemschutzPruefung)
        .filter(AtemschutzPruefung.org_id == org_id)
        .execution_options(include_all_tenants=True)
    )


@router.get("/atemschutz-pruefung", response_class=HTMLResponse)
async def pruefung_liste(
    request: Request,
    db: Session = Depends(get_db),
    geraet_id: int = 0,
    status: str = "",
    von: str = "",
    bis: str = "",
):
    user = _require_login(request)
    q = _pruefungen_query(user.org_id, db)
    if geraet_id:
        q = q.filter(AtemschutzPruefung.geraet_id == geraet_id)
    if von:
        try:
            q = q.filter(AtemschutzPruefung.eingesetzt_am >= date.fromisoformat(von))
        except ValueError:
            pass
    if bis:
        try:
            q = q.filter(AtemschutzPruefung.eingesetzt_am <= date.fromisoformat(bis))
        except ValueError:
            pass
    pruefungen = q.order_by(AtemschutzPruefung.eingesetzt_am.desc(), AtemschutzPruefung.id.desc()).all()
    # Status-Filter in Python statt SQL: "alles_ok" umfasst inzwischen auch die
    # aus Flaschendruck/Warnsignal-Werten berechneten Grenzwert-Checks, die
    # keine eigene DB-Spalte haben (siehe AtemschutzPruefung.alles_ok).
    if status == "nicht_ok":
        pruefungen = [p for p in pruefungen if not p.alles_ok]
    elif status == "ok":
        pruefungen = [p for p in pruefungen if p.alles_ok]
    geraete = (
        db.query(AtemschutzGeraet)
        .filter(AtemschutzGeraet.org_id == user.org_id)
        .execution_options(include_all_tenants=True)
        .order_by(AtemschutzGeraet.nummer)
        .all()
    )
    org_settings = (
        db.query(OrgSettings)
        .filter(OrgSettings.org_id == user.org_id)
        .execution_options(include_all_tenants=True)
        .first()
    )
    return templates.TemplateResponse(request, "atemschutz_pruefung/liste.html", {
        "user": user,
        "pruefungen": pruefungen,
        "geraete": geraete,
        "filter_geraet_id": geraet_id,
        "filter_status": status,
        "filter_von": von,
        "filter_bis": bis,
        "ap_token": org_settings.atemschutz_pruef_token if org_settings else None,
    })


@router.get("/atemschutz-pruefung/geraete", response_class=HTMLResponse)
async def pruefung_geraete_uebersicht(request: Request, db: Session = Depends(get_db)):
    user = _require_login(request)
    geraete = (
        db.query(AtemschutzGeraet)
        .filter(AtemschutzGeraet.org_id == user.org_id)
        .execution_options(include_all_tenants=True)
        .order_by(AtemschutzGeraet.nummer)
        .all()
    )
    letzte: dict[int, AtemschutzPruefung] = {}
    for p in (
        _pruefungen_query(user.org_id, db)
        .order_by(AtemschutzPruefung.eingesetzt_am.desc(), AtemschutzPruefung.id.desc())
        .all()
    ):
        letzte.setdefault(p.geraet_id, p)
    return templates.TemplateResponse(request, "atemschutz_pruefung/geraete_uebersicht.html", {
        "user": user,
        "geraete": geraete,
        "letzte_pruefung": letzte,
    })


@router.get("/atemschutz-pruefung/export.xlsx")
async def pruefung_export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    geraet_id: int = 0,
    status: str = "",
    von: str = "",
    bis: str = "",
):
    user = _require_login(request)
    q = _pruefungen_query(user.org_id, db)
    if geraet_id:
        q = q.filter(AtemschutzPruefung.geraet_id == geraet_id)
    if von:
        try:
            q = q.filter(AtemschutzPruefung.eingesetzt_am >= date.fromisoformat(von))
        except ValueError:
            pass
    if bis:
        try:
            q = q.filter(AtemschutzPruefung.eingesetzt_am <= date.fromisoformat(bis))
        except ValueError:
            pass
    pruefungen = q.order_by(AtemschutzPruefung.eingesetzt_am.desc(), AtemschutzPruefung.id.desc()).all()
    xlsx_bytes = _build_xlsx(pruefungen, org=user.org)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Atemschutzgeraetepruefung.xlsx"'},
    )


def _build_xlsx(pruefungen: list[AtemschutzPruefung], org=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atemschutzgeräteprüfung"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D42225")
    cols = [
        "Datum", "Gerät", "Atemschutzträger", "Ort", "Anlass", "Flasche gewechselt",
        "Flaschendruck (bar)", "Flaschendruck i.O.", "Sichtprüfung",
        "Hochdruck-Dichtprüfung (bar/min)", "Hochdruck-Dichtprüfung i.O.",
        "Warnsignal-Prüfung (bar)", "Warnsignal-Prüfung i.O.", "Einsatzbereit", "Zusatzinfo", "Erfasst am",
    ]
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for ri, p in enumerate(pruefungen, start=2):
        row_data = [
            p.eingesetzt_am.strftime("%d.%m.%Y"),
            p.geraet.anzeige_label if p.geraet else f"#{p.geraet_id}",
            p.traeger_name,
            p.ort_text or "",
            "Einsatz" if p.einsatz_art == "einsatz" else "Übung",
            "Ja" if p.flasche_gewechselt else "Nein",
            p.flaschendruck_bar if p.flaschendruck_bar is not None else "",
            "i.O." if p.flaschendruck_ok else "NICHT i.O.",
            "i.O." if p.sichtpruefung_ok else "NICHT i.O.",
            p.druckabfall_bar if p.druckabfall_bar is not None else "",
            "i.O." if p.hochdruckpruefung_ok else "NICHT i.O.",
            p.rueckzugssignal_bar if p.rueckzugssignal_bar is not None else "",
            "i.O." if p.warnsignal_ok else "NICHT i.O.",
            "i.O." if p.geraet_einsatzbereit_ok else "NICHT i.O.",
            p.defekt_info or "",
            format_local_datetime(p.created_at, org),
        ]
        for ci, val in enumerate(row_data, start=1):
            ws.cell(row=ri, column=ci, value=val)

    ws.auto_filter.ref = ws.dimensions
    widths = [12, 22, 22, 20, 10, 16, 14, 14, 12, 20, 20, 16, 16, 14, 30, 18]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/atemschutz-pruefung/{pruefung_id}", response_class=HTMLResponse)
async def pruefung_detail(request: Request, pruefung_id: int, db: Session = Depends(get_db)):
    user = _require_login(request)
    pruefung = _pruefungen_query(user.org_id, db).filter(AtemschutzPruefung.id == pruefung_id).first()
    if not pruefung:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse(request, "atemschutz_pruefung/detail.html", {
        "user": user,
        "pruefung": pruefung,
    })


@router.get("/atemschutz-pruefung/{pruefung_id}/pdf")
async def pruefung_export_pdf(request: Request, pruefung_id: int, db: Session = Depends(get_db)):
    user = _require_login(request)
    pruefung = _pruefungen_query(user.org_id, db).filter(AtemschutzPruefung.id == pruefung_id).first()
    if not pruefung:
        raise HTTPException(status_code=404)
    from app.services.pdf_service import render_as_pruefung_pdf
    pdf_bytes = render_as_pruefung_pdf([pruefung], user=user, base_url=str(request.base_url))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="Atemschutzpruefung_{pruefung.id}.pdf"'},
    )


@router.post("/atemschutz-pruefung/druck-pdf")
async def pruefung_export_pdf_sammel(
    request: Request,
    db: Session = Depends(get_db),
    ids: list[int] = Form(default=[]),
):
    user = _require_login(request)
    if not ids:
        raise HTTPException(status_code=422, detail="Keine Protokolle ausgewählt")
    pruefungen = (
        _pruefungen_query(user.org_id, db)
        .filter(AtemschutzPruefung.id.in_(ids))
        .order_by(AtemschutzPruefung.eingesetzt_am.desc())
        .all()
    )
    if not pruefungen:
        raise HTTPException(status_code=404)
    from app.services.pdf_service import render_as_pruefung_pdf
    pdf_bytes = render_as_pruefung_pdf(pruefungen, user=user, base_url=str(request.base_url))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="Atemschutzgeraetepruefung_Sammel.pdf"'},
    )
