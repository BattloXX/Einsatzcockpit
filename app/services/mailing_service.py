"""Geschaeftslogik fuer Toggle, Konfiguration, CRUD, CSV und Kampagnen."""

import csv
import io
import json
import re
from datetime import UTC, datetime, timedelta
from sqlalchemy import func

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.core.crypto import decrypt_secret, encrypt_secret
from app.models.mailing import (
    MailingCampaign,
    MailingConfig,
    MailingQueueItem,
    MailingRecipientList,
    MailingRecipientListEntry,
    MailingTemplate,
    MailingSuppressionEntry,
)
from app.models.master import MemberTag


def summarize_filter_json(filter_json: str | None, db: Session | None = None, org_id: int | None = None) -> str:
    """Dynamische Listenfilter kompakt und menschenlesbar beschreiben."""
    try:
        filters = json.loads(filter_json or "{}")
    except (TypeError, ValueError):
        return "Keine Filter"
    if not isinstance(filters, dict):
        return "Keine Filter"

    parts: list[str] = []
    if isinstance(filters.get("active"), bool):
        parts.append("Aktiv" if filters["active"] else "Inaktiv")

    codes = [str(code) for code in filters.get("qualification_codes", []) if code]
    if codes:
        labels = codes
        if db is not None:
            from app.models.master import Qualification

            rows = db.query(Qualification).filter(Qualification.code.in_(codes)).all()
            names = {row.code: row.label for row in rows}
            labels = [names.get(code, code) for code in codes]
        parts.append("Qualifikation: " + ", ".join(labels))

    tag_ids = [int(tag_id) for tag_id in filters.get("tag_ids", []) if str(tag_id).isdigit()]
    if tag_ids:
        labels = [str(tag_id) for tag_id in tag_ids]
        if db is not None:
            from app.models.master import MemberTag

            query = db.query(MemberTag).filter(MemberTag.id.in_(tag_ids))
            if org_id is not None:
                query = query.filter(MemberTag.org_id == org_id)
            rows = query.all()
            names = {row.id: row.name for row in rows}
            labels = [names.get(tag_id, str(tag_id)) for tag_id in tag_ids]
        parts.append("Tag: " + ", ".join(labels))

    if filters.get("member_since_after"):
        parts.append("Mitglied seit: ab " + str(filters["member_since_after"]))
    if filters.get("member_since_before"):
        parts.append("Mitglied seit: bis " + str(filters["member_since_before"]))
    return " + ".join(parts) if parts else "Keine Filter"


def mailing_system_enabled(db):
    from app.models.master import SystemSettings

    row = db.query(SystemSettings).filter(SystemSettings.key == "mailing_module_enabled").first()
    return row is not None and row.value == "true"


def mailing_effective_enabled(org_id, db):
    if org_id is None or not mailing_system_enabled(db):
        return False
    from app.models.master import OrgSettings

    row = db.query(OrgSettings).filter(OrgSettings.org_id == org_id).execution_options(include_all_tenants=True).first()
    return bool(row and row.mailing_module_enabled)


def get_mailing_config(db: Session, org_id: int):
    return db.query(MailingConfig).filter(MailingConfig.org_id == org_id).first()


def save_mailing_config(
    db: Session,
    org_id: int,
    *,
    enabled=False,
    resend_api_key=None,
    resend_webhook_secret=None,
    from_addr=None,
    reply_to_default=None,
    sender_display_name=None,
):
    cfg = get_mailing_config(db, org_id)
    if not cfg:
        cfg = MailingConfig(org_id=org_id)
        db.add(cfg)
    cfg.enabled = bool(enabled)
    cfg.from_addr = (from_addr or "").strip() or None
    cfg.reply_to_default = (reply_to_default or "").strip() or None
    cfg.sender_display_name = (sender_display_name or "").strip() or None
    if resend_api_key is not None and resend_api_key.strip():
        cfg.resend_api_key_enc = encrypt_secret(resend_api_key.strip())
    if resend_webhook_secret is not None and resend_webhook_secret.strip():
        cfg.resend_webhook_secret_enc = encrypt_secret(resend_webhook_secret.strip())
    db.flush()
    return cfg


def mailing_api_key(cfg):
    return decrypt_secret(cfg.resend_api_key_enc) if cfg and cfg.resend_api_key_enc else None

def mailing_webhook_secret(cfg):
    return decrypt_secret(cfg.resend_webhook_secret_enc) if cfg and cfg.resend_webhook_secret_enc else None


def ensure_auto_lists(db: Session, org_id: int):
    for source, name in (("all_members", "Alle Mitglieder"), ("incident_commanders", "Einsatzleiter")):
        if (
            not db.query(MailingRecipientList)
            .filter(MailingRecipientList.org_id == org_id, MailingRecipientList.auto_source == source)
            .first()
        ):
            db.add(MailingRecipientList(org_id=org_id, name=name, kind="auto", auto_source=source))
    db.flush()


def create_template(db, org_id, **data):
    obj = MailingTemplate(org_id=org_id, **data)
    db.add(obj)
    db.flush()
    return obj


def create_recipient_list(db, org_id, **data):
    obj = MailingRecipientList(org_id=org_id, **data)
    db.add(obj)
    db.flush()
    return obj


_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def import_recipients(db: Session, recipient_list: MailingRecipientList, recipients):
    existing = {x[0].lower() for x in db.query(MailingRecipientListEntry.email)
        .filter(MailingRecipientListEntry.list_id == recipient_list.id).all()}
    added = skipped = 0
    for recipient in recipients:
        email = str(recipient.get("email") or "").strip().lower()
        if not _EMAIL.match(email) or email in existing:
            skipped += 1; continue
        db.add(MailingRecipientListEntry(org_id=recipient_list.org_id, list_id=recipient_list.id,
            email=email, display_name=(recipient.get("display_name") or None)))
        existing.add(email); added += 1
    db.flush()
    return {"added": added, "skipped": skipped}


def import_csv(db: Session, recipient_list: MailingRecipientList, content: bytes | str):
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;") if sample.strip() else csv.excel
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    recipients = []
    for row in reader:
        clean = {str(k).strip().lower(): str(v or "").strip() for k, v in row.items() if k is not None}
        email = (clean.get("email") or clean.get("e-mail") or "").lower()
        name = (
            clean.get("display_name")
            or clean.get("name")
            or " ".join(filter(None, [clean.get("vorname"), clean.get("nachname")]))
        )
        recipients.append({"email": email, "display_name": name or None})
    return import_recipients(db, recipient_list, recipients)


def import_xlsx(db: Session, recipient_list: MailingRecipientList, content: bytes):
    """Excel-Empfaenger org-sicher importieren und bestehende Eintraege aktualisieren."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Die Excel-Datei ist leer.") from exc

    headers = {str(value or "").strip().casefold(): index for index, value in enumerate(raw_headers)}
    email_index = next(
        (headers[name] for name in ("e-mail", "email", "e-mail-adresse") if name in headers),
        None,
    )
    if email_index is None:
        raise ValueError("Die Pflichtspalte E-Mail fehlt.")
    name_index = headers.get("name")
    first_name_index = headers.get("vorname")
    last_name_index = headers.get("nachname")
    group_index = headers.get("gruppe")

    existing_entries = {
        entry.email.lower(): entry
        for entry in db.query(MailingRecipientListEntry)
        .filter(
            MailingRecipientListEntry.org_id == recipient_list.org_id,
            MailingRecipientListEntry.list_id == recipient_list.id,
        )
        .all()
    }
    tags_by_name = {
        tag.name.casefold(): tag
        for tag in db.query(MemberTag).filter(MemberTag.org_id == recipient_list.org_id).all()
    }
    added = updated = skipped = 0

    def cell(row, index):
        return str(row[index] or "").strip() if index is not None and index < len(row) else ""

    for row in rows:
        email = cell(row, email_index).lower()
        if not _EMAIL.match(email):
            skipped += 1
            continue
        name = cell(row, name_index)
        if not name:
            name = " ".join(filter(None, [cell(row, first_name_index), cell(row, last_name_index)]))

        row_tags = []
        for raw_tag_name in cell(row, group_index).split(","):
            tag_name = raw_tag_name.strip()
            if not tag_name:
                continue
            key = tag_name.casefold()
            tag = tags_by_name.get(key)
            if tag is None:
                tag = MemberTag(org_id=recipient_list.org_id, name=tag_name)
                db.add(tag)
                db.flush()
                tags_by_name[key] = tag
            if tag not in row_tags:
                row_tags.append(tag)

        entry = existing_entries.get(email)
        if entry is None:
            entry = MailingRecipientListEntry(
                org_id=recipient_list.org_id,
                list_id=recipient_list.id,
                email=email,
                display_name=name or None,
            )
            entry.tags = row_tags
            db.add(entry)
            existing_entries[email] = entry
            added += 1
        else:
            if name:
                entry.display_name = name
            existing_tag_ids = {tag.id for tag in entry.tags}
            entry.tags.extend(tag for tag in row_tags if tag.id not in existing_tag_ids)
            updated += 1
    db.flush()
    return {"added": added, "updated": updated, "skipped": skipped}


def create_campaign(db, org_id, **data):
    obj = MailingCampaign(org_id=org_id, **data)
    db.add(obj)
    db.flush()
    return obj


def queue_campaign(db: Session, campaign: MailingCampaign):
    from app.services.mailing_recipients import resolve_recipient_list_multi

    if campaign.status != "draft":
        return campaign
    recipients = resolve_recipient_list_multi(db, campaign)
    for r in recipients:
        db.add(
            MailingQueueItem(
                org_id=campaign.org_id,
                campaign_id=campaign.id,
                email=r["email"],
                display_name=r.get("display_name"),
                status="queued",
                max_attempts=campaign.max_attempts_override or 3,
            )
        )
    campaign.total_count = len(recipients)
    campaign.status = "queued"
    campaign.queued_at = datetime.now(UTC).replace(tzinfo=None)
    db.flush()
    return campaign

def retry_failed_items(db: Session, campaign: MailingCampaign):
    now = datetime.now(UTC).replace(tzinfo=None)
    db.flush()
    rows = db.query(MailingQueueItem).filter(MailingQueueItem.campaign_id == campaign.id, MailingQueueItem.status == "failed").all()
    for row in rows:
        row.status, row.attempt_count, row.next_attempt_at, row.error_message = "queued", 0, now, None
    if rows:
        campaign.status = "queued"
        campaign.failed_count = max(0, campaign.failed_count - len(rows))
    db.flush()
    return len(rows)

def build_mailing_dashboard_data(db: Session) -> dict:
    campaigns=db.query(MailingCampaign).order_by(MailingCampaign.created_at.desc()).all(); sent=sum(x.sent_count for x in campaigns)
    backlog=db.query(MailingQueueItem).filter(MailingQueueItem.status.in_(["queued","sending"])).count(); cutoff=datetime.now(UTC).replace(tzinfo=None)-timedelta(hours=24)
    failures=db.query(MailingQueueItem).filter(MailingQueueItem.status=="failed",MailingQueueItem.created_at>=cutoff).count()
    days=[(datetime.now(UTC).date()-timedelta(days=x)) for x in range(29,-1,-1)]
    send_rows=db.query(func.date(MailingQueueItem.sent_at),func.count(MailingQueueItem.id)).filter(MailingQueueItem.status.in_(["sent","delivered"]),MailingQueueItem.sent_at>=datetime.combine(days[0],datetime.min.time())).group_by(func.date(MailingQueueItem.sent_at)).all(); send_map={str(k):v for k,v in send_rows}
    categories={"Konfiguration":0,"Rate-Limit":0,"Timeout/Netzwerk":0,"Sonstige":0}
    for (msg,) in db.query(MailingQueueItem.error_message).filter(MailingQueueItem.status=="failed").all():
        low=(msg or "").lower(); key="Konfiguration" if "konfiguration" in low or "auth" in low else "Rate-Limit" if "429" in low or "rate" in low else "Timeout/Netzwerk" if "timeout" in low or "connect" in low else "Sonstige"; categories[key]+=1
    recent=campaigns[:20]
    return {"total_campaigns":len(campaigns),"sent":sent,"open_rate":round(100*sum(x.open_count for x in campaigns)/sent,1) if sent else 0,"click_rate":round(100*sum(x.click_count for x in campaigns)/sent,1) if sent else 0,"backlog":backlog,"failures":failures,"campaign_labels":[x.template.name for x in recent],"campaign_open_rates":[round(100*x.open_count/x.sent_count,1) if x.sent_count else 0 for x in recent],"campaign_click_rates":[round(100*x.click_count/x.sent_count,1) if x.sent_count else 0 for x in recent],"day_labels":[str(x) for x in days],"day_sends":[send_map.get(str(x),0) for x in days],"failure_labels":list(categories),"failure_values":list(categories.values()),"recent_campaigns":campaigns[:5],"campaigns":campaigns}
