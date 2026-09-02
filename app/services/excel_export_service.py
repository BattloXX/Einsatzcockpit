"""Excel-Export des Fahrtenbuchs (openpyxl)."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from app.core.timezones import format_local_datetime
from app.models.fahrtenbuch import Fahrt


def _val(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Ja" if v else "Nein"
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y %H:%M")
    return str(v)


def exportiere_fahrten(fahrten: list[Fahrt], org=None) -> bytes:
    """Erstellt eine Excel-Datei mit allen übergebenen Fahrten.

    Gibt die rohen Bytes zurück (geeignet für StreamingResponse).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl ist nicht installiert. `pip install openpyxl` ausführen.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fahrtenbuch"

    HEADER_FILL = PatternFill("solid", fgColor="2D2D2D")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    headers = [
        "Datum/Zeit", "Fahrzeug", "Kennzeichen",
        "Maschinist", "2. Maschinist",
        "km-Stand", "gefahrene km", "BH-Stand", "BH-Delta",
        "Seilwinde-BH", "Seilwinde-Delta",
        "Zielort", "Zweck", "Zweck-Freitext", "Fahrttyp", "Leitstellennummer",
        "Ausbildner", "Gruppenkommandant", "Einsatzleiter",
        "Schaden", "betriebsfähig", "Schadenbeschreibung",
        "statistikrelevant", "Status",
        "erfasst via", "erfasst von", "Bemerkung",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18

    for f in fahrten:
        fz = f.fahrzeug
        zweck = f.zweck
        zielort_text = (f.zielort.name if f.zielort else None) or f.zielort_freitext or ""
        row = [
            format_local_datetime(f.zeitpunkt, org) if f.zeitpunkt else "",
            fz.code if fz else "",
            fz.kennzeichen if fz and fz.kennzeichen else "",
            f.maschinist_name or "",
            f.maschinist2_name or "",
            f.km_stand_neu if f.km_stand_neu is not None else "",
            f.km_delta if f.km_delta is not None else "",
            str(f.betriebsstunden_neu) if f.betriebsstunden_neu is not None else "",
            str(f.betriebsstunden_delta) if f.betriebsstunden_delta is not None else "",
            str(f.seilwinde_bh_neu) if f.seilwinde_bh_neu is not None else "",
            str(f.seilwinde_bh_delta) if f.seilwinde_bh_delta is not None else "",
            zielort_text,
            zweck.name if zweck else "",
            f.zweck_freitext or "",
            f.fahrttyp.label if f.fahrttyp else "",
            str(f.incident.lis_operation_number or f.incident.nummer or f.incident.id)
            if f.incident_id and f.incident else "",
            f.ausbildner_name or "",
            f.gruppenkommandant_name or "",
            f.einsatzleiter_name or "",
            "Ja" if f.schaden_vorhanden else "Nein",
            ("Ja" if f.schaden_betriebsfaehig else "Nein") if f.schaden_vorhanden else "",
            f.schaden_beschreibung or "",
            "Nein" if f.nicht_statistikrelevant else "Ja",
            f.status.value if f.status else "",
            f.erfasst_via.value if f.erfasst_via else "",
            f.token_label or "",
            f.bemerkung or "",
        ]
        ws.append(row)

    # Spaltenbreiten
    COL_WIDTHS = [16, 12, 12, 22, 22, 10, 12, 10, 10, 12, 14, 22, 22, 24, 10, 12,
                  22, 22, 22, 8, 12, 30, 14, 12, 12, 22, 30]
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportiere_einsatzstatistik(stats, von, bis, org) -> bytes:
    """Erstellt die gefilterte Einsatzstatistik als XLSX-Arbeitsmappe."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl ist nicht installiert. `pip install openpyxl` ausfuehren.")

    wb = openpyxl.Workbook()
    overview = wb.active
    overview.title = "Uebersicht"
    overview.append(["Einsatzstatistik", org.name])
    overview.append(["Zeitraum", f"{von.isoformat()} bis {bis.isoformat()}"])
    overview.append([])
    for label, value in (
        ("Echte Einsaetze", stats.total), ("Uebungen", stats.total_exercises),
        ("Brand", stats.fire_count), ("Technisch", stats.technical_count),
        ("Sonstige", stats.other_count), ("Median Dauer (min)", stats.avg_duration_min),
        ("Median bis erste Fahrzeug-Disposition (min)", stats.avg_time_to_first_vehicle_min),
    ):
        overview.append([label, value if value is not None else ""])

    incidents = wb.create_sheet("Einsaetze")
    incidents.append(["Nummer", "Beginn", "Ende", "Stichwort", "Adresse"])
    for incident in stats.incidents:
        address = " ".join(filter(None, [incident.address_street, incident.address_no, incident.address_city]))
        incidents.append([
            incident.lis_operation_number or incident.nummer or "", format_local_datetime(incident.started_at, org),
            format_local_datetime(incident.closed_at, org), incident.alarm_type_code, address,
        ])

    vehicles = wb.create_sheet("Fahrzeugnutzung")
    vehicles.append(["Fahrzeug", "Bezeichnung", "Zuweisungen", "Kilometer"])
    for item in stats.vehicle_usage:
        vehicles.append([item["code"], item["name"], item["count"], item["km"]])

    fill = PatternFill("solid", fgColor="2D2D2D")
    font = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        ws.freeze_panes = "A2"
        for column in ws.columns:
            width = min(45, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportiere_fahrzeug_links(fahrzeuge: list, org_token: str, base_url: str) -> bytes:
    """Excel mit allen Fahrzeugen inkl. direktem Fahrtenbuch-Link (QR-Deep-Link).

    Der Link öffnet das Erfassungsformular mit vorausgewähltem Fahrzeug:
    {base_url}/f/{org_token}/v/{qr_token}. Fahrzeuge ohne QR-Token erhalten
    keinen Link (Spalte bleibt leer).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl ist nicht installiert. `pip install openpyxl` ausführen.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fahrzeuge"

    HEADER_FILL = PatternFill("solid", fgColor="2D2D2D")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    headers = ["Fahrzeug", "Name", "Kennzeichen", "Typ", "Fahrtenbuch-Link"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18

    base = (base_url or "").rstrip("/")
    for fz in fahrzeuge:
        link = f"{base}/f/{org_token}/v/{fz.qr_token}" if fz.qr_token else ""
        row_idx = ws.max_row + 1
        ws.append([fz.code or "", fz.name or "", fz.kennzeichen or "", fz.type or "", link])
        if link:
            cell = ws.cell(row=row_idx, column=5)
            cell.hyperlink = link
            cell.style = "Hyperlink"

    for i, w in enumerate([16, 26, 14, 20, 70], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportiere_maschinisten_matrix(matrix: dict, org) -> bytes:
    """Erstellt die Maschinisten-Matrix druckfertig als A3-XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        raise RuntimeError("openpyxl ist nicht installiert. `pip install openpyxl` ausführen.")
    from app.services.maschinisten_matrix_service import (
        FARBE_EINSATZ,
        FARBE_STUFE,
        FARBE_UEBUNG,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maschinisten-Matrix"
    ws["A1"] = f"Maschinisten-Matrix {org.name}"
    ws["A2"], ws["B2"] = "M-Stufe", "Name"
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2D2D2D")
    header_font = Font(bold=True, color="FFFFFF")
    col = 3
    for spalte in matrix["spalten"]:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(1, col, spalte["label"])
        ws.cell(2, col, "Üng")
        ws.cell(2, col + 1, "Einsatz")
        col += 2
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    ws.cell(1, col, f"Gesamt {matrix['jahr']}")
    ws.cell(2, col, "Üng")
    ws.cell(2, col + 1, "Einsatz")
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=col + 1):
        for cell in row:
            cell.fill, cell.font, cell.border = header_fill, header_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_no, zeile in enumerate(matrix["zeilen"], 3):
        ws.cell(row_no, 1, f"M{zeile['stufe']}" if zeile["stufe"] else "")
        ws.cell(row_no, 2, zeile["name"])
        if zeile["stufe"]:
            ws.cell(row_no, 1).fill = PatternFill("solid", fgColor=FARBE_STUFE[zeile["stufe"]])
        col = 3
        for spalte in matrix["spalten"]:
            zelle = zeile["zellen"].get(spalte["key"], {})
            for typ, farbe in (("uebung", FARBE_UEBUNG), ("einsatz", FARBE_EINSATZ)):
                wert = zelle.get(typ, 0)
                ws.cell(row_no, col, wert if wert else None)
                ws.cell(row_no, col).fill = PatternFill("solid", fgColor=farbe)
                col += 1
        for typ, farbe in (("uebung", FARBE_UEBUNG), ("einsatz", FARBE_EINSATZ)):
            wert = zeile["summe"][typ]
            ws.cell(row_no, col, wert if wert else None)
            ws.cell(row_no, col).fill = PatternFill("solid", fgColor=farbe)
            col += 1
        for cell in ws[row_no]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.column != 2 else "left")
    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 28
    for col_no in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_no)].width = 8
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
